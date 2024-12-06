import openpyxl as op
import tabulate as tab
from openpyxl.utils import get_column_letter
wb=op.load_workbook("Ventas.xlsx")
ws=wb.active
ws['C2']=0.55
wb.save("mi_libro.xlsx")
wb2=op.load_workbook('mi_libro.xlsx')
ws2=wb2.active
print(ws2['C2'].value)