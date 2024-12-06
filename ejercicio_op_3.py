import openpyxl as op
import tabulate as tab
wb=op.load_workbook("Ventas.xlsx")
ws=wb.active
data=[]
for row in range(1,ws.max_row):
    _row=[]
    for col in ws.iter_cols(1,ws.max_column):
        _row.append(col[row].value)
    data.append(_row)
print(data)